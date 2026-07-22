"""Seed guards and import the 50-student dataset from students.xlsx.

The roll number (e.g. CB.SC.U4CIV24101) is the lookup key. We store it as-is in
`roll_number` for display and a normalised form (CBSCU4CIV24101) in `barcode_id`
so a scan/typed value matches regardless of dots or spacing.

Also writes ../frontend/dataset.js so the web app can show the name instantly
and work offline with the same data.
"""
import os
import re
import json
import openpyxl
from database import SessionLocal, engine, Base
from models import SystemUser, Student
from passlib.context import CryptContext

HERE = os.path.dirname(__file__)
XLSX = os.path.join(HERE, "students.xlsx")
DATASET_JS = os.path.join(HERE, "..", "frontend", "dataset.js")


def norm(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


Base.metadata.create_all(bind=engine)
db = SessionLocal()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Reset
db.query(SystemUser).delete()
db.query(Student).delete()
db.commit()

# Guards
for u, p in [("guard1", "guard123"), ("guard2", "guard123")]:
    db.add(SystemUser(username=u, password_hash=pwd_context.hash(p), role="guard"))

# Students from Excel
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# Wardens — one per distinct hostel (username: warden_<hostel>, password: warden123)
hostels = set()
for r in range(2, ws.max_row + 1):
    h = ws.cell(row=r, column=5).value
    if h:
        hostels.add(str(h).strip())
for h in sorted(hostels):
    db.add(SystemUser(
        username="warden_" + h.lower(),
        password_hash=pwd_context.hash("warden123"),
        role="warden",
        hostel=h,
    ))

js_rows = []
for r in range(2, ws.max_row + 1):
    # "People With Student" (col 6) is intentionally ignored — the number of
    # people in the vehicle is entered at the gate, not taken from the dataset.
    name, roll, phone, place, hostel = [ws.cell(row=r, column=c).value for c in range(1, 6)]
    if not roll:
        continue
    roll = str(roll).strip()
    key = norm(roll)
    db.add(Student(
        barcode_id=key,            # normalised roll, used for matching
        name=str(name).strip() if name else "",
        roll_number=roll,          # display form with dots
        place=str(place).strip() if place else None,
        hostel=str(hostel).strip() if hostel else None,
        phone=str(phone).strip() if phone else None,
    ))
    js_rows.append({
        "key": key,
        "name": str(name).strip() if name else "",
        "roll_number": roll,
        "phone": str(phone).strip() if phone else None,
        "place": str(place).strip() if place else None,
        "hostel": str(hostel).strip() if hostel else None,
    })

db.commit()
count = db.query(Student).count()
db.close()

# Write the frontend dataset (keyed by normalised roll)
with open(DATASET_JS, "w", encoding="utf-8") as f:
    f.write("// Auto-generated from students.xlsx by seed_data.py. Do not edit by hand.\n")
    f.write("window.STUDENT_DATASET = ")
    json.dump(js_rows, f, ensure_ascii=False, indent=0)
    f.write(";\n")

print(f"Done! 2 guards + {len(hostels)} wardens + {count} students imported from students.xlsx")
print(f"Wardens: " + ", ".join("warden_" + h.lower() for h in sorted(hostels)) + " (password: warden123)")
print(f"Wrote {DATASET_JS} ({len(js_rows)} students)")
print("Login: guard1 / guard123")
print(f"Sample roll: {js_rows[0]['roll_number']} -> {js_rows[0]['name']}")
